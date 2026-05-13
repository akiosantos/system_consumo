from email.utils import parseaddr
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

import imaplib
import email
import os
import re
import csv
import shutil

from io import BytesIO
from email.header import decode_header

import pdfplumber

from pypdf import PdfReader, PdfWriter
from openpyxl import load_workbook
from reportlab.pdfgen import canvas

# =========================================================
# CONFIGURAÇÃO
# =========================================================

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# =========================================================
# VARIÁVEIS DE AMBIENTE
# =========================================================

IMAP_SERVER = os.getenv("IMAP_SERVER", "mail.exemplo.com")

SYSTEM_DIR = os.getenv("SYSTEM_DIR", r"C:\Projeto")
BASE_DIR = os.path.join(SYSTEM_DIR, "backend")
FRONT_DIR = os.path.join(SYSTEM_DIR, "frontend")

PLANILHA_ENEL = os.getenv(
    "PLANILHA_ENEL",
    r"C:\Projeto\dados\enel.xlsx"
)

ABA_ENEL = "ENEL"

PLANILHA_SABESP = os.getenv(
    "PLANILHA_SABESP",
    r"C:\Projeto\dados\sabesp.xlsx"
)

ABA_SABESP = "SABESP"

PDF_SABESP_COM_CODIGO = os.path.join(
    BASE_DIR,
    "sabesp_com_codigo.pdf"
)

# =========================================================
# SABESP
# =========================================================

EMAIL_SABESP = os.getenv(
    "EMAIL_SABESP",
    "usuario@empresa.com"
)

SENHA_SABESP = os.getenv(
    "SENHA_SABESP",
    "senha_aqui"
)

REMETENTES_SABESP = [
    "fatura@sabesp.com.br",
    "usuario@empresa.com"
]

PASTA_SABESP = os.path.join(BASE_DIR, "sabesp_pdf")
PASTA_SABESP_SEM_SENHA = os.path.join(
    BASE_DIR,
    "sabesp_pdf_sem_senha"
)

CSV_SABESP = os.path.join(
    BASE_DIR,
    "sabesp_consolidado.csv"
)

PDF_SABESP_COMPLETO = os.path.join(
    BASE_DIR,
    "sabesp_completo.pdf"
)

SENHAS_SABESP = [
    "123",
    "ABC"
]

# =========================================================
# ENEL
# =========================================================

EMAIL_ENEL = os.getenv(
    "EMAIL_ENEL",
    "usuario@empresa.com"
)

SENHA_ENEL = os.getenv(
    "SENHA_ENEL",
    "senha_aqui"
)

REMETENTE_ENEL = [
    "enel.com",
    "usuario@empresa.com"
]

PASTA_ENEL = os.path.join(BASE_DIR, "enel_pdf")

PASTA_ENEL_SEM_SENHA = os.path.join(
    BASE_DIR,
    "enel_pdf_sem_senha"
)

PDF_ENEL_FILTRADO = os.path.join(
    BASE_DIR,
    "enel_filtrado.pdf"
)

PDF_ENEL_COM_CODIGO = os.path.join(
    BASE_DIR,
    "enel_com_codigo.pdf"
)

CSV_ENEL = os.path.join(
    BASE_DIR,
    "enel_consolidado.csv"
)

SENHA_ENEL_PDF = os.getenv(
    "SENHA_ENEL_PDF",
    "12345"
)

# =========================================================
# CRIAÇÃO DAS PASTAS
# =========================================================

os.makedirs(PASTA_SABESP, exist_ok=True)
os.makedirs(PASTA_SABESP_SEM_SENHA, exist_ok=True)

os.makedirs(PASTA_ENEL, exist_ok=True)
os.makedirs(PASTA_ENEL_SEM_SENHA, exist_ok=True)

# =========================================================
# UTILITÁRIOS
# =========================================================

def decodificar(texto):
    partes = decode_header(texto)

    return "".join(
        parte.decode(enc or "utf-8", errors="ignore")
        if isinstance(parte, bytes)
        else parte
        for parte, enc in partes
    )


def normalizar(texto):
    return re.sub(r"\s+", " ", texto.lower()) if texto else ""


def normalizar_instalacao(valor):
    return valor.lstrip("0") if valor else ""


# =========================================================
# PDF
# =========================================================

def tentar_remover_senha(
    caminho_entrada,
    caminho_saida,
    senhas
):
    reader = PdfReader(caminho_entrada)

    if not reader.is_encrypted:
        shutil.copy2(caminho_entrada, caminho_saida)
        return True

    ok = False

    for senha in senhas:
        try:
            if reader.decrypt(senha) != 0:
                ok = True
                break
        except:
            continue

    if not ok:
        return False

    writer = PdfWriter()
    writer.append(reader)

    with open(caminho_saida, "wb") as f:
        writer.write(f)

    return True


def juntar_pdfs(lista_pdfs, pdf_saida):
    writer = PdfWriter()

    for caminho in lista_pdfs:

        if not os.path.exists(caminho):
            print(f"[JUNTAR] Arquivo não encontrado: {caminho}")
            continue

        try:
            reader = PdfReader(caminho)

            for page in reader.pages:
                nova_pagina = writer.add_page(page)
                nova_pagina.mediabox = page.mediabox

                if "/CropBox" in page:
                    nova_pagina.cropbox = page.cropbox

        except Exception as e:
            print(f"[JUNTAR] Erro ao abrir {caminho}: {e}")

    if len(writer.pages) == 0:
        print("[JUNTAR] Nenhum PDF válido encontrado.")
        return False

    with open(pdf_saida, "wb") as f:
        writer.write(f)

    return True


# =========================================================
# EXTRAÇÃO ENEL
# =========================================================

def extrair_instalacao(texto):

    m_mte = re.search(
        r"\b(MTE[A-Z0-9]{5,15})\b",
        texto,
        re.IGNORECASE
    )

    if m_mte:
        return m_mte.group(1).upper()

    numeros = re.findall(r"\b\d{8,12}\b", texto)

    if numeros:
        return normalizar_instalacao(numeros[0])

    return ""


def extrair_referencia(texto, instalacao):

    if not instalacao:
        return ""

    m_ref = re.search(
        r'REFERÊNCIA\s*:\s*([0-9]{2}/[0-9]{4})',
        texto,
        re.IGNORECASE
    )

    if m_ref:
        return m_ref.group(1)

    datas = re.findall(
        r"\b(0[1-9]|1[0-2])/([0-9]{4})\b",
        texto
    )

    if datas:
        return f"{datas[0][0]}/{datas[0][1]}"

    return ""


def extrair_total(texto):

    texto = texto.lower()

    m = re.search(r"r\$\s*([\d.,]+)", texto)

    if m:
        return m.group(1)

    return ""


def extrair_ir(texto):

    texto = texto.lower()

    m = re.search(
        r"irrf?\s*1[,\.]20\s*%",
        texto
    )

    if m:
        return "Encontrado"

    return "0,00"


def extrair_consumo_enel(texto):

    texto = texto.upper()

    m = re.search(
        r"KWH\s+([\d.,]+)",
        texto
    )

    if m:
        return m.group(1)

    return ""


# =========================================================
# EXTRAÇÃO SABESP
# =========================================================

def extrair_fornecimento_sabesp(texto):

    m = re.search(r"\b(\d{9,16})\b", texto)

    if m:
        return m.group(1)

    return ""


def extrair_consumo_sabesp(texto):

    texto = re.sub(r"\s+", " ", texto)

    m = re.search(
        r"\d{2}/\d{2}/\d{2}\s+\d{1,6}\s+(\d{1,6})",
        texto
    )

    if m:
        return m.group(1)

    return ""


# =========================================================
# PLANILHAS
# =========================================================

def carregar_mapa_instalacao_codigo():

    wb = load_workbook(
        PLANILHA_ENEL,
        data_only=True
    )

    ws = wb[ABA_ENEL]

    mapa = {}

    for linha in ws.iter_rows(min_row=2):

        codigo = linha[1].value
        instalacao = linha[3].value

        if codigo and instalacao:
            mapa[
                normalizar_instalacao(str(instalacao))
            ] = str(codigo)

    return mapa


def carregar_mapa_fornecimento_codigo_sabesp():

    wb = load_workbook(
        PLANILHA_SABESP,
        data_only=True
    )

    ws = wb[ABA_SABESP]

    mapa = {}

    for linha in ws.iter_rows(min_row=2):

        codigo = linha[1].value
        fornecimento = linha[4].value

        if codigo and fornecimento:
            mapa[str(fornecimento)] = str(codigo)

    return mapa


# =========================================================
# ENDPOINT TESTE
# =========================================================

@app.get("/status")
def status():
    return {
        "status": "online"
    }


# =========================================================
# FRONTEND
# =========================================================

app.mount(
    "/",
    StaticFiles(directory=FRONT_DIR, html=True),
    name="frontend"
)

